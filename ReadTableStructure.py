from tkinter import *
import tkinter.filedialog
import openpyxl
from openpyxl import workbook

def pickFileA():
    filename = tkinter.filedialog.askopenfilename(filetypes = [('XPO','xpo')])

    if filename !='':
        entry1.delete(0,END)
        entry1.insert(10,filename)


def readTableStructure():
    myTableStructure  = TableStructure()
    readFile(entry1,myTableStructure)
    writeExcel(myTableStructure)


def readFile(entryObj,myTableStructure):
    fileSource = open(entryObj.get())
    GroupListStart = False
    while True:
        line = fileSource.readline()
        if(not line):
            break
        if line[0:9] == "  TABLE #":
            myTableStructure.TableName = line[9:].strip()
        if line[0:13] == "      FIELD #":
            myTableStructure.FieldList.add(line[13:].strip())
        if line[0:13] == "    ENDFIELDS":
            break
    fileSource.close()

def writeExcel(myTableStructure):
    sourceFile = str(entry1.get())
    targetFile = sourceFile.replace(".xpo",".xlsx")
    myPos = sourceFile.rfind("/")
    titleName = sourceFile[myPos+1:].replace("Table_","").replace(".xpo","")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = titleName

    sheet.cell(1,1,"TABLE")
    sheet.cell(1,2,myTableStructure.TableName)
    sheet.cell(2,1,"FIELDS")
    sheet.cell(2,2,"Description")

    i = 3
    for field in myTableStructure.FieldList:
        sheet.cell(i,1,field)
        i+=1
    
    workbook.save(targetFile)

    textbox1 = Text(myWindow)
    textbox1.config(wrap=WORD)
    textbox1.insert(END,"The structure for {0} has been saved to {1}.".format(titleName,targetFile))
    textbox1.grid(row = 3,columnspan = 24)


class TableStructure:
    TableName = ''
    FieldList = set()
    

#初始化TK()
myWindow = Tk()
#设置标题
myWindow.title('Read table structure from xpo')
#设置窗口大小
#myWindow.geometry('380x300')
width = 380
height = 300
#获取屏幕尺寸以计算布局参数，使窗口居屏幕中央
screenwidth = myWindow.winfo_screenwidth()
screenheight = myWindow.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width,height,(screenwidth -width)/2,(screenheight - height)/2)
myWindow.geometry(alignstr)

#放置控件
Label(myWindow,text = "Pick up xpo file:").grid(row = 0)
entry1 = Entry(myWindow)
entry1.grid(row = 0, column = 1)
button1 = Button(myWindow,text = "...",command = pickFileA)
button1.grid(row = 0, column = 2)

button3 = Button(myWindow, text = "Read table", command = readTableStructure)
button3.grid(row = 2,column = 2)
#进入消息循环
myWindow.mainloop()